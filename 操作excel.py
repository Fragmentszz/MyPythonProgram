import openpyxl
from  openpyxl import  Workbook
import asyncio
import re
import cn2an

def min2hmin(minz):
    h=minz//60
    minz=minz % 60
    if h==0:
        return '{}min'.format(minz)
    elif minz==0:
        return '{}h'.format(h)
    else:
        return '{}h{}min'.format(h,minz)

def time(time,time0):
    #请以月份.日期的形式输入time0
    listz=time0.split('-')
    rq1=listz[0].replace('.','-')
    rq2=listz[1].replace('.','-')
    
    if rq1<=time[5:-9] and time[5:-9]<=rq2:
        return True
    else:
        return False

def swicth(sc:str):
    sc=sc.replace("小时半","小时30分钟").replace("个",'')
    sc=cn2an.transform(sc,"cn2an")
    sc=sc.replace('小时','h').replace('分钟','min')
    ph='(\d+\.?\d?)h'
    pmin='(\d+)min'
    hz=0
    minzz=0
    if sc.find('h') >0:
        h_list=re.findall(ph,sc)
        for h in h_list:
            hz=hz+eval(h)
    else:
        h=0
    if sc.find('min')>0:
        min_list=re.findall(pmin,sc)
        for minz in min_list:
            minzz=eval(minz)+minzz
    else:
        minz=0
    return (int(hz*60+minzz))

def bj(tuplez):
    return tuplez[1]

def zqsport(time0,sort:str,path=None):
    dic={}
    workbook=openpyxl.load_workbook('C://Users/ASUS/Desktop/zz.xlsx')
    zaoqi=workbook[sort]
    max_row=zaoqi.max_row
    dic2={}
    for gyd,dyg,xh,zy in zip(zaoqi['B'],zaoqi['C'],zaoqi['D'],zaoqi['E']):
        if time(time=str(gyd.value),time0=time0) :
            dic[dyg.value]=dic.get(dyg.value,0)+1
            try:
                cs = dic2[dyg.value]
            except:
                dic2[dyg.value] = xh.value + zy.value
    li=list(dic.items())
    li.sort(key=bj,reverse=True)
    li2=[]
    for z in li:
        tp=(dic2[z[0]],)
        z=z+tp
        li2.append(z)
    return li2




    
def ydxx(time0,sort:str,path=None):
    dic={}
    dic1={}
    workbook=openpyxl.load_workbook('C://Users/ASUS/Desktop/zz.xlsx')
    yuedu=workbook[sort]
    i=0
    dic2={}
    for sj,xm,sc,xh,zy in zip(yuedu['B'],yuedu['C'],yuedu['F'],yuedu['D'],yuedu['E']):
        if time(time=str(sj.value),time0=time0):
            dic[xm.value]=dic.get(xm.value,0)+swicth(sc.value)
            try:
                cs = dic2[xm.value]
            except:
                dic2[xm.value] = xh.value + zy.value

    li=list(dic.items())
    li.sort(key=bj,reverse=True)
    li2=[]
    for z in li:
        tp=(dic2[z[0]],)
        z=z+ tp
        li2.append(z)
    return li2


def count(time0,sort,path=None):
    dic={}
    workbook=openpyxl.load_workbook('C://Users/ASUS/Desktop/zz.xlsx')
    zaoqi=workbook[sort]
    max_row=zaoqi.max_row
    for gyd,dyg in zip(zaoqi['B'],zaoqi['C']):
        if time(time=str(gyd.value),time0=time0) :
            dic[dyg.value]=dic.get(dyg.value,0)+1
    return dic


def cz1(li_list:list,sheet) :
    i=2
    for li in li_list:
        sheet["A{}".format(i)] = li[0]
        sheet["D{}".format(i)] = li[1]
        str = li[2]
        sheet["B{}".format(i)] = str[9:]
        sheet["C{}".format(i)] = str[:9]
        i+=1

def cz2(li_list:list,dic:dict,sheet):
    i=2
    for li in li_list:
        sheet["A{}".format(i)] = li[0]
        sheet["D{}".format(i)] = min2hmin(li[1])
        sheet["B{}".format(i)] = li[2][9:]
        sheet["C{}".format(i)] = li[2][:9]
        sheet["E{}".format(i)]=dic[li[0]]
        i += 1


def result():
    workbook = openpyxl.load_workbook('C://Users/ASUS/Desktop/result.xlsx')
    for i in range(4):
        if i == 0 :
            li_list = zqsport("01.08-02.20","早起")
            sheet = workbook["早起"]
            cz1(li_list , sheet)
        if i ==1:
            li_list = zqsport("01.08-02.20","运动")
            sheet = workbook["运动"]
            cz1(li_list, sheet)
        if i ==2:
            li_list = ydxx("01.01-02.20","学习")
            sheet = workbook["学习"]
            dic=count("01.01-02.20","学习")
            cz2(li_list,dic,sheet)
        if i ==3:
            li_list = ydxx("01.01-02.20", "阅读")
            sheet = workbook["阅读"]
            dic = count("01.01-02.20", "阅读")
            cz2(li_list, dic, sheet)
    workbook.save('结果.xlsx')

result()
